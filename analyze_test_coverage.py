"""Analyze test coverage between Excel test cases and automated tests"""
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Read Excel
df = pd.read_excel('API-Testcases.xlsx')

# Add columns
if 'Automation Status' not in df.columns:
    df['Automation Status'] = ''
if 'Automated Test Location' not in df.columns:
    df['Automated Test Location'] = ''
if 'Coverage Notes' not in df.columns:
    df['Coverage Notes'] = ''

# Test mapping (abbreviated - full mapping in separate file)
test_mapping = {
    'TC-PCI-001': ('test_schemathesis_comprehensive.py', 'test_api_schema_compliance', 'Schemathesis'),
    'TC-PCI-002': ('test_schemathesis_comprehensive.py', 'TestPublishCourseInventory::test_valid_course_filters', 'Explicit'),
    # ... (mapping continues in full script)
}

# Update dataframe
for idx, row in df.iterrows():
    test_id = row['Test Case ID']
    if test_id in test_mapping:
        file_name, test_name, notes = test_mapping[test_id]
        df.at[idx, 'Automation Status'] = '✅ AUTOMATED'
        df.at[idx, 'Automated Test Location'] = f'{file_name}::{test_name}'
        df.at[idx, 'Coverage Notes'] = notes
    else:
        df.at[idx, 'Automation Status'] = '❌ NOT AUTOMATED'

# Save
output_file = 'API-Testcases-Updated.xlsx'
df.to_excel(output_file, index=False)

# Color code
wb = load_workbook(output_file)
ws = wb.active
green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

status_col = None
for col in range(1, ws.max_column + 1):
    if ws.cell(1, col).value == 'Automation Status':
        status_col = col
        break

if status_col:
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row, status_col)
        if '✅' in str(cell.value):
            cell.fill = green_fill
        elif '❌' in str(cell.value):
            cell.fill = red_fill

wb.save(output_file)

# Summary
total = len(df)
automated = len(df[df['Automation Status'].str.contains('✅', na=False)])
coverage = (automated / total * 100) if total > 0 else 0

print(f"\nCoverage: {automated}/{total} ({coverage:.1f}%)")
print(f"Output: {output_file}")
